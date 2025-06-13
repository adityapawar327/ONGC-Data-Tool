import pandas as pd
from fuzzywuzzy import fuzz
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import streamlit as st
import re


class FuzzySearcher:
    def __init__(self):
        self.low_threshold = 60
        self.medium_threshold = 75
        self.high_threshold = 90
        self.current_query = ""

        # Define synonyms for normalization
        self.synonyms_map = {
            "input": ["input", "i/o", "io", "ip", "input/output"],
            "output": ["output", "out", "o/p", "op", "output/input"],
            "error": ["error", "err", "fault", "fail"],
            "user": ["user", "usr", "client", "operator"],
        }

        self.reverse_synonyms = {}
        for key, syns in self.synonyms_map.items():
            for syn in syns:
                self.reverse_synonyms[syn] = key

    def normalize_text(self, text):
        text = text.lower()
        text = re.sub(r'[^a-z0-9\s]', '', text)
        words = text.split()
        normalized_words = []
        for w in words:
            normalized_words.append(self.reverse_synonyms.get(w, w))
        return " ".join(normalized_words)

    def get_match_level(self, similarity, text):
        # Special terms we're looking for
        data_types = ["SEGY", "DATA", "RAW DATA", "WATER BOTTOM", "CUBE"]
        
        # Convert both to uppercase for comparison
        text_upper = text.upper()
        query_upper = self.current_query.upper()
        
        # If user is searching for one of our special terms
        if any(term in query_upper for term in data_types):
            # Exact match gets green
            if text_upper == query_upper:
                return "high"
            # Partial match gets yellow if above threshold
            elif similarity >= self.low_threshold:
                return "medium"
        # For other searches, just use similarity threshold
        elif similarity >= self.low_threshold:
            return "medium"
        return None

    def search_dataframe(self, df: pd.DataFrame, query: str, columns=None):
        # Store the current query for use in get_match_level
        self.current_query = query
        
        if df.empty or not query:
            return {"matches": [], "columns": list(df.columns), "count": 0, "cells": []}

        norm_query = self.normalize_text(query)
        matches = []
        cell_refs = []

        search_columns = df.columns if columns is None else [col for col in columns if col in df.columns]

        for row_idx, row in df.iterrows():
            for col_idx, col_name in enumerate(search_columns):
                value = row[col_name]
                if pd.isnull(value):
                    continue

                try:
                    cell_text = str(value)
                    norm_cell_text = self.normalize_text(cell_text)

                    similarity = fuzz.token_set_ratio(norm_query, norm_cell_text)
                    match_level = self.get_match_level(similarity, cell_text)
                    
                    if match_level:
                        col_index = df.columns.get_loc(col_name)
                        matches.append((row_idx, col_index, similarity, match_level))
                        cell_refs.append(f"{get_column_letter(col_index + 1)}{row_idx + 2}")
                except Exception:
                    continue

        return {
            "matches": matches,
            "columns": list(df.columns),
            "count": len(matches),
            "cells": cell_refs,
        }

def get_highlight_color(match_level):
    return {
        "medium": "background-color: #FFD700",    # Dark golden yellow
        "high": "background-color: #32CD32"       # Lime green
    }[match_level]

def highlight_cells(df: pd.DataFrame, matches: list):
    matched_cells = {(m[0], m[1]): m[3] for m in matches}  # (row, col): match_level

    def highlight_cell(val, row, col):
        if (row, col) in matched_cells:
            return get_highlight_color(matched_cells[(row, col)])
        return ""

    styled_df = df.style.applymap(lambda v: "", subset=df.columns).apply(
        lambda x: [highlight_cell(val, x.name, idx) for idx, val in enumerate(x)],
        axis=1
    )
    return styled_df


def create_excel_with_highlights(df: pd.DataFrame, matches: list):
    wb = Workbook()
    ws = wb.active    # Define colors for different match levels
    colors = {
        "medium": "FFD700",    # Dark golden yellow
        "high": "32CD32"       # Lime green
    }

    ws.append(list(df.columns))

    for i, row in df.iterrows():
        for j, val in enumerate(row):
            cell_val = "" if pd.isnull(val) else val
            ws.cell(row=i + 2, column=j + 1).value = cell_val

    for match in matches:
        row_idx, col_idx, _, match_level = match
        cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
        cell.fill = PatternFill(
            start_color=colors[match_level],
            end_color=colors[match_level],
            fill_type="solid"
        )

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer


def highlight_single_cell(df: pd.DataFrame, match):
    row_idx, col_idx, _, match_level = match
    
    def highlight_cell(val, row, col):
        if (row, col) == (row_idx, col_idx):
            return get_highlight_color(match_level)
        return ""

    styled_df = df.style.applymap(lambda v: "", subset=df.columns).apply(
        lambda x: [highlight_cell(val, x.name, idx) for idx, val in enumerate(x)],
        axis=1
    )
    return styled_df


def create_excel_single_cell(df: pd.DataFrame, match):
    return create_excel_with_highlights(df, [match])


def read_file(uploaded_file):
    try:
        if uploaded_file.name.endswith(".csv"):
            return pd.read_csv(uploaded_file)
        else:
            return pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(f"Error reading file '{uploaded_file.name}': {e}")
        return None


def export_matches_to_excel(df, matches_by_row, selected_indices, df_name):
    import xlsxwriter
    from io import BytesIO
    import base64

    # Create Excel file in memory
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()    # Define formats for different match levels
    formats = {
        "medium": workbook.add_format({'bg_color': '#FFD700'}),    # Dark golden yellow
        "high": workbook.add_format({'bg_color': '#32CD32'})       # Lime green
    }

    # Write headers
    for col_idx, col_name in enumerate(df.columns):
        worksheet.write(0, col_idx, col_name)

    # Write data with highlighting
    excel_row = 1
    for row_idx in selected_indices:
        row_data = df.iloc[row_idx]
        
        # Write all cells in the row
        for col_idx, value in enumerate(row_data):
            cell_format = None
            
            # Check if this cell has a match and apply appropriate format
            if row_idx in matches_by_row:
                for match in matches_by_row[row_idx]:
                    if match["col_idx"] == col_idx:
                        cell_format = formats[match["match_level"]]
                        break
            
            worksheet.write(excel_row, col_idx, value, cell_format)
        
        excel_row += 1    # Add legend
    legend_row = excel_row + 2
    worksheet.write(legend_row, 0, "Match Levels:")
    worksheet.write(legend_row + 1, 0, "Exact Data Type Match", formats["high"])
    worksheet.write(legend_row + 2, 0, "Similar Match (60% or higher)", formats["medium"])

    workbook.close()

    # Generate download link
    output.seek(0)
    b64 = base64.b64encode(output.read()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="fuzzy_search_results_{df_name}.xlsx">Download Excel file</a>'
    st.markdown(href, unsafe_allow_html=True)


def fuzzy_search_ui():
    st.header("🔍 Multi-file Fuzzy Search Excel/CSV Data")
    st.markdown("""
    This tool searches through your files and highlights matches:
    - 🟨 Dark Yellow: Similar matches (60% or higher similarity)
    - 🟢 Green: Exact matches for data types (SEGY, DATA, WATER BOTTOM, etc.)
    """)

    uploaded_files = st.file_uploader(
        "Upload one or more Excel/CSV files",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=True,
        key="multi_search"
    )
    search_term = st.text_input("Enter search term")

    if uploaded_files and search_term:
        searcher = FuzzySearcher()
        all_results = []
        total_matches = 0
        file_dfs = {}
        dropdown_options = []

        for uploaded_file in uploaded_files:
            df = read_file(uploaded_file)
            if df is None:
                continue

            file_dfs[uploaded_file.name] = df
            results = searcher.search_dataframe(df, search_term)
            total_matches += results["count"]

            for idx, cell in enumerate(results["cells"]):
                match = results["matches"][idx]
                similarity = match[2]
                match_level = match[3]
                level_emoji = {"low": "🟡", "medium": "🟠", "high": "🟢"}[match_level]
                
                dropdown_options.append({
                    "label": f"{level_emoji} {uploaded_file.name} - {cell} (Sim: {similarity}%)",
                    "file": uploaded_file.name,
                    "cell": cell,
                    "match": match,
                    "results": results,
                })

            all_results.append({"filename": uploaded_file.name, "df": df, "results": results})

        st.write(f"### Total matches found across files: {total_matches}")

        # Rest of the UI code remains the same
        if total_matches > 0:
            option = st.radio(
                "Highlight options:",
                ("Highlight all matching cells", "Highlight a single selected match")
            )

            if option == "Highlight all matching cells":
                for file_result in all_results:
                    st.write(f"#### File: {file_result['filename']}")
                    df = file_result["df"]
                    matches = file_result["results"]["matches"]
                    if matches:
                        styled_df = highlight_cells(df, matches)
                        st.dataframe(styled_df, use_container_width=True)

                        excel_bytes = create_excel_with_highlights(df, matches)
                        st.download_button(
                            label=f"📥 Download Highlighted Excel for {file_result['filename']}",
                            data=excel_bytes,
                            file_name=f"highlighted_{file_result['filename']}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.write("No matches in this file.")

            else:  # Highlight single selected match
                selected_match_label = st.selectbox(
                    "Select a matched cell to highlight",
                    options=[opt["label"] for opt in dropdown_options]
                )

                selected_match = next((opt for opt in dropdown_options if opt["label"] == selected_match_label), None)
                if selected_match:
                    file_name = selected_match["file"]
                    df = file_dfs[file_name]
                    match = selected_match["match"]
                    cell_address = selected_match["cell"]

                    styled_df = highlight_single_cell(df, match)
                    st.write(f"#### Preview highlighting cell {cell_address} in file: {file_name}")
                    st.dataframe(styled_df, use_container_width=True)

                    excel_bytes = create_excel_single_cell(df, match)
                    st.download_button(
                        label=f"📥 Download Excel with single highlight ({cell_address})",
                        data=excel_bytes,
                        file_name=f"highlighted_single_{file_name}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    st.markdown(f"➡️ **Navigate to cell {cell_address} in Excel** by clicking that cell after opening the downloaded file.")

        else:
            st.info("No matches found in uploaded files.")
